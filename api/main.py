import hashlib
import io
import json
import os
from datetime import datetime, timedelta, timezone
from typing import Optional

import anthropic
import duckdb
import openpyxl
import pandas as pd
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from jose import JWTError, jwt
from pydantic import BaseModel, field_validator
from supabase import Client, create_client

load_dotenv()

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "datasets")

DATASETS = [
    {
        "filename": "novaretail_sales_q1_2024.csv",
        "name": "novaretail_sales_q1_2024.csv",
        "description": "NovaRetail Inc. Q1 2024 sales data — orders, products, categories, revenue, regions, and stores.",
    },
]

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL     = os.environ["SUPABASE_URL"]
SUPABASE_KEY     = os.environ["SUPABASE_KEY"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
JWT_SECRET       = os.environ["JWT_SECRET"]
JWT_ALGORITHM    = "HS256"
JWT_EXPIRE_HOURS = 24

# ── Clients ───────────────────────────────────────────────────────────────────
_supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
_ai = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ── Job paths (mirrors core/job_paths.py — no st dependency here) ─────────────
JOB_PATHS = {
    "junior_data_analyst": {
        "id": "junior_data_analyst",
        "title": "Junior Data Analyst",
        "company": "NovaRetail Inc.",
        "industry": "Retail & E-commerce",
        "manager": "Sophie Chen",
        "manager_title": "Senior Data Analyst",
        "stack": ["SQL", "Python", "Excel", "Power BI"],
        "duration_weeks": 24,
        "responsibilities": [
            "Clean and transform sales and inventory datasets",
            "Create weekly reports for the operations team",
            "Analyze customer behavior trends",
            "Maintain and update Power BI dashboards",
            "Handle ad-hoc data requests from the business team",
        ],
        "schema": """
NovaRetail Inc. — Database Schema (use these exact table and column names in all SQL tasks):

Table: sales
  - order_id (INTEGER) — unique order identifier
  - product_name (TEXT) — name of the product sold
  - category (TEXT) — product category (e.g. Electronics, Apparel, Home, Food)
  - units_sold (INTEGER) — number of units in the order
  - unit_price (DECIMAL) — price per unit
  - total_revenue (DECIMAL) — units_sold * unit_price
  - sale_date (DATE) — date of the transaction (format: YYYY-MM-DD)
  - region (TEXT) — sales region (North, South, East, West, International)
  - store_id (INTEGER) — store where the sale occurred

Table: inventory
  - product_id (INTEGER) — unique product identifier
  - product_name (TEXT) — name of the product
  - category (TEXT) — product category
  - stock_level (INTEGER) — current units in stock
  - reorder_point (INTEGER) — minimum stock before reorder is triggered
  - last_restocked (DATE) — date of last restock

Table: customers
  - customer_id (INTEGER) — unique customer identifier
  - full_name (TEXT) — customer full name
  - email (TEXT) — customer email
  - region (TEXT) — customer region
  - signup_date (DATE) — date they joined NovaRetail
  - total_orders (INTEGER) — lifetime number of orders
  - total_spent (DECIMAL) — lifetime spend
""",
    }
}

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="Simployee API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Auth helpers ──────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    """SHA-256 — matches app.py and core/database.py."""
    return hashlib.sha256(password.encode()).hexdigest()

def create_token(user_id: int, username: str) -> str:
    payload = {
        "sub": str(user_id),
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token.",
            headers={"WWW-Authenticate": "Bearer"},
        )

_bearer = HTTPBearer()

def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(_bearer),
) -> dict:
    payload = decode_token(credentials.credentials)
    result = (
        _supabase.table("users")
        .select("id, username, full_name, created_at")
        .eq("id", int(payload["sub"]))
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found.")
    return result.data[0]

# ── Progress helpers ──────────────────────────────────────────────────────────
def get_or_create_progress(user_id: int) -> dict:
    """Return user_progress row, creating one with the default job path if absent."""
    result = _supabase.table("user_progress").select("*").eq("user_id", user_id).execute()
    if result.data:
        return result.data[0]
    insert = _supabase.table("user_progress").insert({
        "user_id":     user_id,
        "job_path_id": "junior_data_analyst",
    }).execute()
    return insert.data[0]

# ── Task helpers ──────────────────────────────────────────────────────────────
def get_task_owned_by(task_id: int, user_id: int) -> dict:
    result = (
        _supabase.table("tasks")
        .select("*")
        .eq("id", task_id)
        .eq("user_id", user_id)
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found.")
    return result.data[0]

# ── File reading (mirrors 1_Tasks.py logic) ───────────────────────────────────
async def read_uploaded_file(file: UploadFile) -> str:
    content = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
            return df.to_string(max_rows=100)
        elif name.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(content))
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = [
                    "\t".join(str(c) if c is not None else "" for c in row)
                    for row in ws.iter_rows(max_row=50, values_only=True)
                ]
                parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
            return "\n\n".join(parts)
        else:
            return content.decode("utf-8", errors="ignore")
    except Exception as e:
        return f"Could not read file: {e}"

# ── AI: generate tasks (mirrors core/ai_engine.py) ────────────────────────────
def _generate_weekly_tasks(job_path: dict, week_number: int) -> list[dict]:
    prompt = f"""You are a simulation engine for a job training platform called Simployee.

The user has been hired as a {job_path['title']} at {job_path['company']},
a company in the {job_path['industry']} industry.

Their manager is {job_path['manager']} ({job_path['manager_title']}).

Key responsibilities:
{chr(10).join(f"- {r}" for r in job_path['responsibilities'])}

Required stack: {', '.join(job_path['stack'])}

Company database schema (always reference exact table/column names in SQL tasks):
{job_path.get('schema', 'No schema provided.')}

Generate 3 realistic work tasks for Week {week_number} of their simulated job.
Each task should feel like a real ticket a junior analyst would receive at work.

Rules:
- Tasks must be practical and doable (the user will actually attempt them)
- Increase complexity gradually based on the week number
- Tasks must relate to the company context (retail, sales, inventory, customers)
- When tasks involve SQL, always mention the exact table and column names from the schema above
- Week 1-4: Basic tasks (data cleaning, simple queries, basic reports)
- Week 5-12: Intermediate tasks (analysis, visualization, Python scripts)
- Week 13-24: Advanced tasks (dashboards, automation, insights presentations)
- For SQL tasks: the user works against novaretail_sales_q1_2024.csv, available for download on the platform. Never ask for results from a live database. The deliverable is the SQL query itself, optionally with results shown from running it on the CSV using pandas or SQLite.

Respond ONLY with a valid JSON array. No explanation, no markdown, no extra text.
Format:
[
  {{
    "title": "Short task title",
    "description": "2-3 sentences describing the task with business context, as if written by the manager in Slack. If the task involves SQL, mention the exact table and columns to use.",
    "deliverable": "Exactly what the user must submit (e.g. SQL query, Python script, Excel file, written analysis)"
  }},
  {{"title": "...", "description": "...", "deliverable": "..."}},
  {{"title": "...", "description": "...", "deliverable": "..."}}
]"""

    message = _ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    return json.loads(raw)

# ── AI: Sophie chat (chat-based feedback per task) ───────────────────────────
def _sophie_chat_response(
    task: dict,
    job_path: dict,
    history: list[dict],
    new_user_message: str,
    file_content: Optional[str] = None,
    file_name: Optional[str] = None,
) -> str:
    system = f"""You are Sophie Chen, {job_path['manager_title']} at {job_path['company']}.
You are in an ongoing conversation with your junior analyst about this task:

Task: {task['title']}
Description: {task['description']}
Expected deliverable: {task['deliverable']}

Guidelines:
- Write exactly like a real manager messaging on Slack — direct, warm, human. No corporate speak.
- NEVER start with "Great job!" or "Excellent work!" or any generic praise opener.
- On the first message: give specific feedback on what they submitted. Be honest.
- On follow-up messages: continue naturally, answer questions, give guidance.
- If the work is fundamentally correct, be positive and give ONE concrete improvement tip only.
- If there's a real problem, call out ONE main issue — the most important one. Don't list every flaw.
- For SQL tasks: the user works against novaretail_sales_q1_2024.csv. The deliverable is the SQL query itself — results shown from running it on the CSV with pandas or SQLite are a bonus.
- NEVER suggest using Notion, Jira, Confluence, Trello, Asana, GitHub Issues, Linear, monday.com, or any external tracking system. If they need to document something, tell them to include it in their message or a plain text file.
- NEVER reference external tools that don't exist in the simulation: no Slack channels, shared Google Docs, or external links.
- Keep it 100-200 words.
- Sound like a real person, not an AI. No bullet point lists.
- End with a follow-up question or next step."""

    messages = []
    for m in history:
        role    = m.get("role")
        content = m.get("content", "")
        if role in ("user", "assistant") and content:
            messages.append({"role": role, "content": content})

    if file_content and new_user_message.strip():
        user_content = f"{new_user_message}\n\n[File: {file_name}]\n{file_content[:3000]}"
    elif file_content:
        user_content = f"[File: {file_name}]\n{file_content[:3000]}"
    else:
        user_content = new_user_message

    messages.append({"role": "user", "content": user_content})

    response = _ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=400,
        system=system,
        messages=messages,
    )
    return response.content[0].text.strip()


# ── AI: evaluate submission (mirrors core/ai_engine.py) ───────────────────────
def _evaluate_submission(
    task: dict,
    submission_text: str,
    job_path: dict,
    file_content: Optional[str] = None,
    file_name: Optional[str] = None,
) -> str:
    if file_content and submission_text.strip():
        submission_context = f"""The analyst submitted both a file and a written comment.

File uploaded: {file_name}
File content (first 3000 chars):
{file_content[:3000]}

Written comment:
{submission_text}"""
    elif file_content:
        submission_context = f"""The analyst submitted a file only.

File uploaded: {file_name}
File content (first 3000 chars):
{file_content[:3000]}"""
    elif submission_text.strip():
        submission_context = f"""The analyst submitted a written response only (no file attached).

Written response:
{submission_text}"""
    else:
        submission_context = "The analyst submitted nothing meaningful."

    prompt = f"""You are Sophie Chen, a {job_path['manager_title']} at {job_path['company']}.
You are reviewing work submitted by your junior analyst.

Task assigned:
Title: {task['title']}
Description: {task['description']}
Expected deliverable: {task['deliverable']}

What was submitted:
{submission_context}

Instructions for your response:
- Write exactly like a real manager messaging their junior on Slack — direct, human, no corporate speak
- NEVER start with "Great job!" or "Excellent work!" or any generic praise opener
- If the deliverable was supposed to be a file (Excel, CSV, script) and they only sent text, call it out directly and ask for the actual file. Example tone: "Hey, I appreciate the explanation but I actually need the file itself here, not just a description of what you did."
- If they submitted a file but it's missing something, point it out specifically
- If the work is fundamentally correct, the tone must be positive — acknowledge what they did well, then give exactly one concrete improvement tip. Do not pile on multiple issues.
- If the work has a real problem, call out ONE main issue only — the most important one. Do not list every flaw you notice.
- NEVER suggest using Notion, Jira, Confluence, Trello, Asana, GitHub Issues, Linear, monday.com, or any external tracking system. If the analyst identifies a data issue or needs to document something, tell them to include it directly in their submission text or in a plain text file — nothing else.
- Keep it 150-250 words
- Sound like a real person, not an AI. No bullet point lists in your response.
- End with either a follow-up question or a next step, like a real manager would"""

    message = _ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text.strip()

# ── Request schemas ───────────────────────────────────────────────────────────
class RegisterRequest(BaseModel):
    full_name: str
    username:  str
    password:  str

    @field_validator("full_name", "username")
    @classmethod
    def not_empty(cls, v: str) -> str:
        if not v.strip():
            raise ValueError("Field cannot be blank.")
        return v.strip()

    @field_validator("password")
    @classmethod
    def min_length(cls, v: str) -> str:
        if len(v) < 6:
            raise ValueError("Password must be at least 6 characters.")
        return v

class LoginRequest(BaseModel):
    username: str
    password: str

class SelectJobPathRequest(BaseModel):
    job_path_id: str

class SqlRunRequest(BaseModel):
    query: str

    @field_validator("query")
    @classmethod
    def validate_query(cls, v: str) -> str:
        stripped = v.strip()
        if not stripped.upper().lstrip("(").startswith("SELECT"):
            raise ValueError("Only SELECT queries are allowed.")
        blocked = ["read_csv", "read_parquet", "read_json", "read_text", "glob(", "httpfs", "http://", "https://"]
        lower = stripped.lower()
        for pat in blocked:
            if pat in lower:
                raise ValueError(f"Not allowed in queries: '{pat}'")
        return stripped

class InterviewMessageRequest(BaseModel):
    job_path_id: str
    history: list[dict]

# ── Shared response shape ─────────────────────────────────────────────────────
def _user_payload(user: dict) -> dict:
    return {"id": user["id"], "username": user["username"], "full_name": user["full_name"]}

# ═══════════════════════════════════════════════════════════════════════════════
# Auth routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/auth/register", status_code=status.HTTP_201_CREATED)
def register(body: RegisterRequest):
    try:
        result = _supabase.table("users").insert({
            "username":  body.username,
            "password":  hash_password(body.password),
            "full_name": body.full_name,
        }).execute()
    except Exception:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Username already taken.")

    user  = result.data[0]
    token = create_token(user["id"], user["username"])
    return {"token": token, "user": _user_payload(user)}


@app.post("/auth/login")
def login(body: LoginRequest):
    result = _supabase.table("users").select("*").eq("username", body.username).execute()

    invalid = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Invalid username or password.",
    )
    if not result.data:
        raise invalid

    user = result.data[0]
    if user["password"] != hash_password(body.password):
        raise invalid

    token = create_token(user["id"], user["username"])
    return {"token": token, "user": _user_payload(user)}


@app.get("/auth/me")
def me(current_user: dict = Depends(get_current_user)):
    return current_user

# ═══════════════════════════════════════════════════════════════════════════════
# User progress routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/users/me/progress")
def get_progress(current_user: dict = Depends(get_current_user)):
    result = _supabase.table("user_progress").select("*").eq("user_id", current_user["id"]).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="No job path selected.")
    progress = result.data[0]
    job_path = JOB_PATHS.get(progress["job_path_id"], {})
    return {
        "current_week":  progress["current_week"],
        "job_path_id":   progress["job_path_id"],
        "started_at":    progress.get("started_at"),
        "job_path": {
            "title":          job_path.get("title"),
            "company":        job_path.get("company"),
            "duration_weeks": job_path.get("duration_weeks"),
        },
    }


@app.post("/users/me/progress", status_code=status.HTTP_201_CREATED)
def select_job_path(
    body: SelectJobPathRequest,
    current_user: dict = Depends(get_current_user),
):
    if body.job_path_id not in JOB_PATHS:
        raise HTTPException(status_code=400, detail="Invalid job path ID.")

    existing = (
        _supabase.table("user_progress")
        .select("id")
        .eq("user_id", current_user["id"])
        .execute()
    )
    if existing.data:
        raise HTTPException(status_code=409, detail="Job path already selected.")

    result = _supabase.table("user_progress").insert({
        "user_id":     current_user["id"],
        "job_path_id": body.job_path_id,
    }).execute()
    return result.data[0]


@app.post("/users/me/advance")
def advance_week(current_user: dict = Depends(get_current_user)):
    progress = get_or_create_progress(current_user["id"])
    week     = progress["current_week"]
    job_path = JOB_PATHS.get(progress["job_path_id"], {})

    # Verify all tasks for this week are reviewed
    tasks_result = (
        _supabase.table("tasks")
        .select("status")
        .eq("user_id", current_user["id"])
        .eq("week", week)
        .execute()
    )
    tasks = tasks_result.data or []

    if not tasks:
        raise HTTPException(status_code=400, detail="Generate and complete tasks before advancing.")
    if any(t["status"] != "reviewed" for t in tasks):
        raise HTTPException(status_code=400, detail="Complete all tasks before advancing to the next week.")

    max_weeks = job_path.get("duration_weeks", 24)
    if week >= max_weeks:
        raise HTTPException(status_code=400, detail="You have completed the full program!")

    result = (
        _supabase.table("user_progress")
        .update({"current_week": week + 1})
        .eq("user_id", current_user["id"])
        .execute()
    )
    return {"current_week": result.data[0]["current_week"]}

# ═══════════════════════════════════════════════════════════════════════════════
# Task routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/tasks")
def list_tasks(week: int, current_user: dict = Depends(get_current_user)):
    result = (
        _supabase.table("tasks")
        .select("*")
        .eq("user_id", current_user["id"])
        .eq("week", week)
        .order("id")
        .execute()
    )
    return result.data or []


@app.post("/tasks/generate", status_code=status.HTTP_201_CREATED)
def generate_tasks(current_user: dict = Depends(get_current_user)):
    progress = get_or_create_progress(current_user["id"])
    week     = progress["current_week"]
    job_path = JOB_PATHS.get(progress["job_path_id"])

    if not job_path:
        raise HTTPException(status_code=400, detail="Invalid job path.")

    # Prevent re-generation if tasks already exist for this week
    existing = (
        _supabase.table("tasks")
        .select("id")
        .eq("user_id", current_user["id"])
        .eq("week", week)
        .execute()
    )
    if existing.data:
        raise HTTPException(status_code=409, detail="Tasks for this week already exist.")

    try:
        generated = _generate_weekly_tasks(job_path, week)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {e}")

    rows = [
        {
            "user_id":     current_user["id"],
            "week":        week,
            "title":       t["title"],
            "description": t["description"],
            "deliverable": t["deliverable"],
        }
        for t in generated
    ]
    result = _supabase.table("tasks").insert(rows).execute()
    return result.data


@app.get("/tasks/{task_id}/messages")
def get_task_messages(task_id: int, current_user: dict = Depends(get_current_user)):
    get_task_owned_by(task_id, current_user["id"])
    result = (
        _supabase.table("task_messages")
        .select("id, task_id, role, content, created_at")
        .eq("task_id", task_id)
        .order("created_at")
        .execute()
    )
    return result.data or []


@app.post("/tasks/{task_id}/message", status_code=status.HTTP_201_CREATED)
async def send_task_message(
    task_id:      int,
    message:      str                  = Form(""),
    file:         Optional[UploadFile] = File(None),
    current_user: dict                 = Depends(get_current_user),
):
    task = get_task_owned_by(task_id, current_user["id"])

    if not message.strip() and not file:
        raise HTTPException(status_code=400, detail="Send something — text, code, or a file.")

    file_content: Optional[str] = None
    file_name:    Optional[str] = None
    if file and file.filename:
        file_content = await read_uploaded_file(file)
        file_name    = file.filename

    if file_name and message.strip():
        user_content_db = f"[File: {file_name}]\n\n{message}"
    elif file_name:
        user_content_db = f"[File: {file_name}]"
    else:
        user_content_db = message

    history_res = (
        _supabase.table("task_messages")
        .select("role, content")
        .eq("task_id", task_id)
        .order("created_at")
        .execute()
    )
    history = history_res.data or []

    progress = get_or_create_progress(current_user["id"])
    job_path = JOB_PATHS.get(progress["job_path_id"], {})

    try:
        sophie_reply = _sophie_chat_response(
            task, job_path, history, message,
            file_content=file_content, file_name=file_name,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI response failed: {e}")

    _supabase.table("task_messages").insert([
        {"task_id": task_id, "role": "user",      "content": user_content_db},
        {"task_id": task_id, "role": "assistant",  "content": sophie_reply},
    ]).execute()

    if task["status"] == "pending":
        _supabase.table("tasks").update({"status": "reviewed"}).eq("id", task_id).execute()

    return {"role": "assistant", "content": sophie_reply}

# ═══════════════════════════════════════════════════════════════════════════════
# Dataset routes
# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# Interview routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/interview/message")
def interview_message(
    body: InterviewMessageRequest,
    current_user: dict = Depends(get_current_user),
):
    if body.job_path_id not in JOB_PATHS:
        raise HTTPException(status_code=400, detail="Invalid job path.")

    job_path = JOB_PATHS[body.job_path_id]
    user_message_count = sum(1 for m in body.history if m.get("role") == "user")
    is_final = user_message_count >= 3

    system = f"""You are Sophie Chen, {job_path['manager_title']} at {job_path['company']}.
You are doing a quick screening interview for a {job_path['title']} position.

Follow this structure strictly:
- Turn 1 (no user messages yet): Introduce yourself warmly, say you're excited about the role, then ask what drew them to data analysis and why they're applying.
- Turn 2 (after 1 user answer): Ask a beginner-friendly practical question about their experience with {job_path['stack'][0]} or working with data in general.
- Turn 3 (after 2 user answers): Ask how they handle receiving feedback on their work or a time they had to learn something new quickly.
- Turn 4 (after 3 user answers): Give warm, honest final feedback. Reference 1-2 specific things they said. Highlight 2 genuine strengths and 1 area to grow. Close with: "Welcome to the team — can't wait to get started with you!"

Style rules:
- Write like a real person messaging on Slack — casual, warm, human. No corporate language.
- No bullet points, no numbered questions, no headers.
- Keep each message under 120 words."""

    messages = [{"role": "user", "content": "Ready."}]
    for m in body.history:
        role = m.get("role")
        content = m.get("content", "")
        if role in ("user", "assistant") and content:
            messages.append({"role": role, "content": content})

    response = _ai.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=300,
        system=system,
        messages=messages,
    )

    return {
        "message": response.content[0].text.strip(),
        "is_final": is_final,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Dataset routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/sql/run")
def run_sql(body: SqlRunRequest, current_user: dict = Depends(get_current_user)):
    csv_path = os.path.join(DATA_DIR, "novaretail_sales_q1_2024.csv").replace("\\", "/")
    if not os.path.isfile(csv_path):
        raise HTTPException(status_code=500, detail="Dataset file not found on server.")

    try:
        conn = duckdb.connect()
        conn.execute(f"CREATE VIEW sales AS SELECT * FROM read_csv_auto('{csv_path}')")
        rel = conn.execute(
            f"SELECT * FROM ({body.query.rstrip(';')}) __q LIMIT 100"
        )
        columns = [desc[0] for desc in rel.description]
        rows = rel.fetchall()
    except duckdb.Error as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    finally:
        try:
            conn.close()
        except Exception:
            pass

    def _safe(v):
        if v is None:
            return None
        if isinstance(v, (bool, int, float, str)):
            return v
        return str(v)

    return {
        "columns": columns,
        "rows":    [[_safe(v) for v in row] for row in rows],
        "row_count": len(rows),
    }


@app.get("/datasets")
def list_datasets(request: Request):
    base = str(request.base_url).rstrip("/")
    return [
        {
            "name": d["name"],
            "description": d["description"],
            "download_url": f"{base}/datasets/{d['filename']}/download",
        }
        for d in DATASETS
    ]


@app.get("/datasets/{filename}/download")
def download_dataset(filename: str):
    known = {d["filename"] for d in DATASETS}
    if filename not in known:
        raise HTTPException(status_code=404, detail="Dataset not found.")
    path = os.path.join(DATA_DIR, filename)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="File not found on server.")
    return FileResponse(path, filename=filename, media_type="text/csv")
